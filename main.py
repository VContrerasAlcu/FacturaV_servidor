from fastapi import FastAPI, Depends, HTTPException, status, File, UploadFile, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from typing import List, Optional
import re
import io
from datetime import datetime
from PIL import Image
import logging

from config import settings
from database import init_db, get_user_by_email, save_user, verify_password, hash_password
from auth import (
    create_access_token, 
    get_current_user, 
    generate_verification_code, 
    store_verification_code, 
    validate_verification_code,
    get_verification_data,
    remove_verification_code
)
from models import (
    UserCreate, 
    UserResponse, 
    Token, 
    VerificationRequest, 
    PasswordResetRequest,
    ProcessResponse
)
from email_sender import send_verification_code, send_email
from image_processor import process_image
from excel_generator import generate_excel
from contextlib import asynccontextmanager

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Inicializar base de datos
    try:
        init_db()
        logger.info("Base de datos inicializada correctamente")
    except Exception as e:
        logger.error(f"Error inicializando base de datos: {e}")
    yield

# Inicializar aplicación FastAPI
app = FastAPI(title="FacturaV API", version="1.0.0", lifespan=lifespan)

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Módulo para compresión de imágenes
async def compress_image(file: UploadFile, max_size_mb: int = 4, quality: int = 85) -> UploadFile:
    """
    Comprime una imagen si excede el tamaño máximo permitido
    """
    try:
        # Leer el contenido del archivo
        content = await file.read()
        
        # Verificar si necesita compresión (4MB límite de Azure DI)
        if len(content) <= max_size_mb * 1024 * 1024:
            # Resetear el archivo para lectura posterior
            file.file.seek(0)
            return file
        
        logger.info(f"Comprimiendo imagen {file.filename} de {len(content)/1024/1024:.2f}MB")
        
        # Abrir imagen con PIL
        image = Image.open(io.BytesIO(content))
        
        # Convertir a RGB si es necesario (para JPEG)
        if image.mode in ('RGBA', 'P', 'LA'):
            image = image.convert('RGB')
        
        # Calcular factor de compresión
        original_size_mb = len(content) / 1024 / 1024
        compression_ratio = (max_size_mb * 0.9) / original_size_mb  # Usar 90% del límite
        new_quality = max(40, int(quality * compression_ratio))  # Calidad mínima 40%
        
        # Comprimir imagen
        output = io.BytesIO()
        image.save(output, format='JPEG', quality=new_quality, optimize=True)
        compressed_content = output.getvalue()
        
        logger.info(f"Imagen comprimida: {len(compressed_content)/1024/1024:.2f}MB (calidad: {new_quality}%)")
        
        # Crear nuevo UploadFile con el contenido comprimido
        compressed_file = UploadFile(
            filename=f"compressed_{file.filename}",
            file=io.BytesIO(compressed_content),
            content_type='image/jpeg'
        )
        
        return compressed_file
        
    except Exception as e:
        logger.error(f"Error comprimiendo imagen {file.filename}: {e}")
        # En caso de error, devolver el archivo original
        file.file.seek(0)
        return file

# Rutas de autenticación
@app.post("/api/login", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        user = get_user_by_email(form_data.username)
        
        if not user or not verify_password(form_data.password, user['password']):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Credenciales incorrectas",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        if not user['activo']:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Cuenta desactivada",
            )
        
        access_token = create_access_token(data={"sub": user['email']})
        return {"access_token": access_token, "token_type": "bearer"}
    except Exception as e:
        logger.error(f"Error en login: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

@app.post("/api/register")
async def register(user_data: UserCreate, background_tasks: BackgroundTasks):
    try:
        # Validar formato de email
        if not re.match(r"[^@]+@[^@]+\.[^@]+", user_data.email):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato de email inválido"
            )
        
        # Verificar si el usuario ya existe
        if get_user_by_email(user_data.email):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="El usuario ya existe"
            )
        
        # Generar código de verificación
        code = generate_verification_code()
        store_verification_code(
            user_data.email, 
            code, 
            user_data.model_dump(),
            "register"
        )
        
        # Enviar código por email (en background)
        background_tasks.add_task(send_verification_code, user_data.email, code)
        
        return {"message": "Código de verificación enviado"}
    except Exception as e:
        logger.error(f"Error en registro: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

@app.post("/api/verify-code", response_model=Token)
async def verify_code(verification_request: VerificationRequest):
    try:
        if not validate_verification_code(verification_request.email, verification_request.code):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Código inválido o expirado"
            )
        
        # Obtener datos del usuario
        verification_data = get_verification_data(verification_request.email)
        
        if verification_data['action'] == "register":
            # Crear usuario
            user_data = verification_data['user_data']
            hashed_password = hash_password(user_data['password'])
            
            user = {
                'email': user_data['email'],
                'password': hashed_password,
                'nombre': user_data.get('nombre'),
                'dni_cif': user_data.get('dni_cif'),
                'direccion': user_data.get('direccion'),
                'activo': True
            }
            
            save_user(user)
        
        # Eliminar código de verificación
        remove_verification_code(verification_request.email)
        
        # Crear token JWT
        access_token = create_access_token(data={"sub": verification_request.email})
        return {"access_token": access_token, "token_type": "bearer"}
    except Exception as e:
        logger.error(f"Error en verificación: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

@app.post("/api/forgot-password")
async def forgot_password(email: str, background_tasks: BackgroundTasks):
    try:
        user = get_user_by_email(email)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Usuario no encontrado"
            )
        
        # Generar código de verificación
        code = generate_verification_code()
        store_verification_code(
            email, 
            code, 
            None,
            "password_reset"
        )
        
        # Enviar código por email (en background)
        background_tasks.add_task(send_verification_code, email, code)
        
        return {"message": "Código de verificación enviado"}
    except Exception as e:
        logger.error(f"Error en forgot-password: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

@app.post("/api/reset-password")
async def reset_password(password_request: PasswordResetRequest):
    try:
        if not validate_verification_code(password_request.email, password_request.code):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Código inválido o expirado"
            )
        
        # Actualizar contraseña
        hashed_password = hash_password(password_request.new_password)
        user = get_user_by_email(password_request.email)
        
        if user:
            user['password'] = hashed_password
            save_user(user)
        
        # Eliminar código de verificación
        remove_verification_code(password_request.email)
        
        return {"message": "Contraseña actualizada correctamente"}
    except Exception as e:
        logger.error(f"Error en reset-password: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

@app.get("/api/me", response_model=UserResponse)
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    try:
        return current_user
    except Exception as e:
        logger.error(f"Error obteniendo usuario: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno del servidor"
        )

# Ruta para procesar una sola factura (mantener para compatibilidad)
@app.post("/api/upload-invoice", response_model=ProcessResponse)
async def upload_invoice(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    try:
        logger.info(f"📄 Procesando factura individual: {file.filename}")
        
        # Validar tipo de archivo
        if not file.content_type.startswith('image/'):
            return ProcessResponse(
                message="El archivo debe ser una imagen",
                success=False
            )
        
        # Comprimir imagen si es necesario
        compressed_file = await compress_image(file)
        
        # Procesar imagen con Azure Document Intelligence
        processed_data = process_image(compressed_file)
        
        if not processed_data or not processed_data[0]:
            return ProcessResponse(
                message="No se pudieron extraer datos de la factura",
                success=False
            )
        
        # Agregar información de origen a los datos
        for data_item in processed_data:
            data_item['archivo_origen'] = file.filename
            data_item['timestamp_procesamiento'] = datetime.now().isoformat()
        
        # Generar archivo Excel
        excel_file = generate_excel(processed_data)
        
        # Enviar por email (en background)
        background_tasks.add_task(
            send_email,
            current_user['email'], 
            "Factura procesada - FacturaV", 
            "Adjunto encontrará el archivo Excel con los datos de su factura procesada.",
            excel_file, 
            "factura.xlsx"
        )
        
        return ProcessResponse(
            message="Imagen procesada y enviada al email",
            success=True
        )
        
    except Exception as e:
        logger.error(f"❌ Error procesando factura individual: {e}")
        return ProcessResponse(
            message=f"Error procesando imagen: {str(e)}",
            success=False
        )

# Endpoint para procesar múltiples facturas - CORREGIDO
@app.post("/api/upload-invoices", response_model=ProcessResponse)
async def upload_invoices(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    try:
        # DEBUG: Información inicial
        logger.info(f"🎯 INICIO PROCESAMIENTO MÚLTIPLE")
        logger.info(f"📦 Número de archivos recibidos: {len(files)}")
        
        # Validar que se hayan subido archivos
        if not files:
            logger.warning("❌ No se han subido archivos")
            return ProcessResponse(
                message="No se han subido archivos",
                success=False
            )
        
        # Validar número máximo de archivos
        max_files = 10
        if len(files) > max_files:
            logger.warning(f"❌ Demasiados archivos: {len(files)} (máximo {max_files})")
            return ProcessResponse(
                message=f"Máximo {max_files} archivos permitidos",
                success=False
            )
        
        # Validar tipos de archivo
        invalid_files = []
        valid_files = []
        
        for i, file in enumerate(files):
            logger.info(f"📄 Archivo {i+1}: {file.filename} - Tipo: {file.content_type}")
            if file.content_type and file.content_type.startswith('image/'):
                valid_files.append(file)
            else:
                invalid_files.append(file.filename)
        
        if invalid_files:
            logger.warning(f"📛 Archivos inválidos rechazados: {invalid_files}")
        
        if not valid_files:
            logger.error("❌ Ningún archivo válido encontrado")
            return ProcessResponse(
                message="Ninguno de los archivos es una imagen válida",
                success=False
            )
        
        logger.info(f"✅ Archivos válidos para procesar: {len(valid_files)}")
        
        # Procesar cada imagen
        all_processed_data = []
        processed_count = 0
        failed_count = 0
        processing_details = []

        for i, file in enumerate(valid_files):
            try:
                logger.info(f"🔄 Procesando archivo {i+1}/{len(valid_files)}: {file.filename}")
                
                # Comprimir imagen antes de procesar
                compressed_file = await compress_image(file)
                logger.info(f"✅ Imagen {file.filename} comprimida exitosamente")
                
                # Procesar imagen con Azure Document Intelligence
                processed_data = process_image(compressed_file)
                
                if processed_data and len(processed_data) > 0:
                    # Agregar información del archivo a CADA elemento de datos
                    for data_item in processed_data:
                        data_item['archivo_origen'] = file.filename
                        data_item['numero_factura'] = f"{i+1}"
                        data_item['indice_procesamiento'] = i + 1
                        data_item['timestamp_procesamiento'] = datetime.now().isoformat()
                    
                    # EXTENDER la lista, no hacer append
                    all_processed_data.extend(processed_data)
                    processed_count += 1
                    processing_details.append(f"✓ {file.filename}: {len(processed_data)} elementos procesados")
                    logger.info(f"✅ Archivo {file.filename} procesado exitosamente - {len(processed_data)} elementos")
                else:
                    failed_count += 1
                    processing_details.append(f"✗ {file.filename}: no se pudieron extraer datos")
                    logger.warning(f"⚠️ No se pudieron extraer datos del archivo: {file.filename}")
                    
            except Exception as e:
                failed_count += 1
                error_msg = str(e)
                if "too large" in error_msg.lower():
                    error_msg = "imagen demasiado grande (se intentó comprimir pero aún excede el límite)"
                processing_details.append(f"✗ {file.filename}: error - {error_msg}")
                logger.error(f"❌ Error procesando archivo {file.filename}: {e}")
        
        # VERIFICAR resultados del procesamiento
        logger.info(f"📊 RESULTADO DEL PROCESAMIENTO:")
        logger.info(f"   • Elementos procesados: {len(all_processed_data)}")
        logger.info(f"   • Archivos exitosos: {processed_count}")
        logger.info(f"   • Archivos fallidos: {failed_count}")
        logger.info(f"   • Total archivos: {len(valid_files)}")
        
        # Verificar archivos únicos procesados
        archivos_unicos = set()
        for data in all_processed_data:
            if 'archivo_origen' in data:
                archivos_unicos.add(data['archivo_origen'])
        
        logger.info(f"📁 Archivos únicos con datos: {len(archivos_unicos)}")
        logger.info(f"📂 Lista: {list(archivos_unicos)}")
        
        # Verificar si se procesó al menos una factura
        if not all_processed_data:
            logger.error("❌ No se pudo procesar ninguna factura")
            return ProcessResponse(
                message="No se pudieron procesar ninguna de las facturas",
                success=False,
                details=processing_details
            )
        
        # Generar archivo Excel con TODAS las facturas procesadas
        logger.info(f"📊 Generando Excel con {len(all_processed_data)} elementos...")
        excel_file = generate_excel(all_processed_data)
        
        if not excel_file:
            logger.error("❌ No se pudo generar el archivo Excel")
            return ProcessResponse(
                message="Error generando el archivo de resultados",
                success=False,
                details=processing_details
            )
        
        # Verificar el Excel generado
        try:
            excel_content = excel_file.getvalue()
            logger.info(f"✅ Excel generado exitosamente - Tamaño: {len(excel_content)} bytes")
        except Exception as e:
            logger.error(f"❌ Error verificando Excel: {e}")
        
        # Preparar mensaje de resultado
        result_message = f"Procesamiento completado: {processed_count} factura(s) procesada(s) correctamente"
        if failed_count > 0:
            result_message += f", {failed_count} factura(s) fallaron"
        
        # Preparar contenido del email
        email_subject = f"Facturas procesadas ({processed_count}) - FacturaV"
        
        email_content = f"""
        <h3>Procesamiento de facturas completado</h3>
        <p><strong>Resultado:</strong> {result_message}</p>
        <p><strong>Total de archivos procesados:</strong> {len(valid_files)}</p>
        <p><strong>Elementos extraídos:</strong> {len(all_processed_data)}</p>
        <p><strong>Archivos únicos procesados:</strong> {len(archivos_unicos)}</p>
        <p><strong>Fecha de procesamiento:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <h4>Detalles del procesamiento:</h4>
        <ul>
        """
        
        for detail in processing_details:
            email_content += f"<li>{detail}</li>"
        
        email_content += """
        </ul>
        <p>Adjunto encontrará el archivo Excel con los datos de todas las facturas procesadas correctamente.</p>
        <p><strong>El archivo contiene {len(archivos_unicos)} hojas, una para cada factura procesada.</strong></p>
        """
        
        # Enviar por email (en background)
        background_tasks.add_task(
            send_email,
            current_user['email'], 
            email_subject, 
            email_content,
            excel_file, 
            f"facturas_procesadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        logger.info("✅ Email programado para envío en background")
        
        return ProcessResponse(
            message=result_message,
            success=True,
            details=processing_details,
            processed_count=processed_count,
            failed_count=failed_count,
            total_files=len(valid_files),
            unique_files_processed=len(archivos_unicos),
            total_elements=len(all_processed_data)
        )
        
    except Exception as e:
        logger.error(f"💥 Error crítico procesando múltiples facturas: {e}")
        return ProcessResponse(
            message=f"Error procesando las imágenes: {str(e)}",
            success=False
        )

# Ruta para verificar estado del servidor
@app.get("/")
async def root():
    return {"message": "FacturaV API está funcionando correctamente"}

# Ruta de health check
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "FacturaV API",
        "features": {
            "single_upload": True,
            "multiple_upload": True,
            "image_compression": True,
            "max_file_size_mb": 4,
            "max_files_per_request": 10
        }
    }

# Ruta para obtener información del sistema
@app.get("/api/system-info")
async def system_info(current_user: dict = Depends(get_current_user)):
    return {
        "service": "FacturaV API",
        "version": "1.0.0",
        "timestamp": datetime.now().isoformat(),
        "user": current_user['email'],
        "features": {
            "single_upload": True,
            "multiple_upload": True,
            "image_compression": True,
            "max_files": 10,
            "max_file_size": "4MB"
        }
    }

# Endpoint de prueba para verificar que la API está funcionando
@app.get("/api/test")
async def test_endpoint():
    return {
        "message": "API funcionando correctamente",
        "timestamp": datetime.now().isoformat(),
        "endpoints_available": [
            "/api/upload-invoice",
            "/api/upload-invoices", 
            "/api/login",
            "/api/register",
            "/api/verify-code"
        ]
    }

# Endpoint para probar compresión de imágenes
@app.post("/api/test-compression")
async def test_compression(file: UploadFile = File(...)):
    try:
        original_size = len(await file.read())
        file.file.seek(0)
        
        compressed_file = await compress_image(file)
        compressed_size = len(await compressed_file.read())
        
        return {
            "original_size_mb": round(original_size / 1024 / 1024, 2),
            "compressed_size_mb": round(compressed_size / 1024 / 1024, 2),
            "compression_ratio": round((original_size - compressed_size) / original_size * 100, 1),
            "filename": file.filename
        }
    except Exception as e:
        return {"error": str(e)}

# Endpoint para debug de múltiples archivos
@app.post("/api/debug-upload")
async def debug_upload(files: List[UploadFile] = File(...)):
    """
    Endpoint especial para debug que muestra información detallada de los archivos recibidos
    """
    file_info = []
    
    for i, file in enumerate(files):
        content = await file.read()
        file_info.append({
            "index": i + 1,
            "filename": file.filename,
            "content_type": file.content_type,
            "size_bytes": len(content),
            "size_mb": round(len(content) / 1024 / 1024, 2)
        })
        # Resetear el archivo para lectura posterior
        await file.seek(0)
    
    return {
        "total_files": len(files),
        "files_received": file_info,
        "timestamp": datetime.now().isoformat()
    }

# Endpoint para test específico de generación de Excel con múltiples facturas
@app.post("/api/debug-excel")
async def debug_excel_generation(files: List[UploadFile] = File(...)):
    """
    Endpoint para debug específico del Excel
    """
    try:
        all_processed_data = []
        
        for i, file in enumerate(files):
            logger.info(f"🔍 Procesando {file.filename}...")
            
            compressed_file = await compress_image(file)
            processed_data = process_image(compressed_file)
            
            if processed_data:
                for data_item in processed_data:
                    data_item['archivo_origen'] = file.filename
                    data_item['indice'] = i + 1
                    data_item['timestamp'] = datetime.now().isoformat()
                
                all_processed_data.extend(processed_data)
                logger.info(f"✅ {file.filename} → {len(processed_data)} elementos")
            else:
                logger.warning(f"⚠️ {file.filename} → 0 elementos")
        
        # Generar Excel
        excel_file = generate_excel(all_processed_data)
        
        if excel_file:
            # Leer el Excel generado para analizarlo
            import io
            import openpyxl
            
            excel_content = excel_file.getvalue()
            workbook = openpyxl.load_workbook(io.BytesIO(excel_content))
            
            sheet_info = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info.append({
                    'nombre': sheet_name,
                    'filas': sheet.max_row,
                    'columnas': sheet.max_column
                })
            
            return {
                "success": True,
                "archivos_procesados": len(files),
                "elementos_totales": len(all_processed_data),
                "hojas_excel": sheet_info,
                "tamaño_bytes": len(excel_content),
                "mensaje": f"Excel con {len(sheet_info)} hojas generado"
            }
        else:
            return {
                "success": False,
                "mensaje": "Error generando Excel"
            }
            
    except Exception as e:
        logger.error(f"❌ Error en debug-excel: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.get("/api/check-sendgrid")
async def check_sendgrid():
    """Endpoint para verificar la configuración de SendGrid"""
    return {
        "sendgrid_configured": bool(settings.SENDGRID_API_KEY),
        "from_email": settings.FROM_EMAIL,
        "api_key_length": len(settings.SENDGRID_API_KEY) if settings.SENDGRID_API_KEY else 0
    }

# Agrega estos endpoints en main.py para debugging

@app.get("/api/debug/email-config")
async def debug_email_config():
    """Verificar configuración de email"""
    return {
        "sendgrid_configured": bool(settings.SENDGRID_API_KEY),
        "from_email": settings.FROM_EMAIL,
        "api_key_prefix": settings.SENDGRID_API_KEY[:10] + "..." if settings.SENDGRID_API_KEY else "No configurada"
    }

@app.post("/api/debug/test-email-simple")
async def test_email_simple(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Test simple de email sin adjuntos"""
    try:
        from email_sender import send_email
        
        # Email simple sin adjuntos
        success = send_email(
            current_user['email'],
            "TEST Simple - FacturaV",
            "<h1>Test Simple</h1><p>Si recibes esto, el email básico funciona.</p>"
        )
        
        return {
            "success": success,
            "message": "Email de prueba enviado" if success else "Error enviando email",
            "user_email": current_user['email']
        }
        
    except Exception as e:
        logger.error(f"❌ Error en test-email-simple: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/debug/test-email-with-attachment")
async def test_email_with_attachment(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Test de email con archivo adjunto"""
    try:
        from email_sender import send_email_with_file
        from io import BytesIO
        
        # Crear un archivo de prueba simple
        test_content = b"Este es un archivo de prueba generado por FacturaV"
        test_file = BytesIO(test_content)
        
        success = send_email_with_file(
            current_user['email'],
            "TEST Con Adjunto - FacturaV",
            "<h1>Test con Adjunto</h1><p>Si recibes esto con el archivo, todo funciona.</p>",
            test_file,
            "test.txt"
        )
        
        return {
            "success": success,
            "message": "Email con adjunto enviado" if success else "Error enviando email con adjunto",
            "user_email": current_user['email'],
            "file_size": len(test_content)
        }
        
    except Exception as e:
        logger.error(f"❌ Error en test-email-with-attachment: {e}")
        return {"success": False, "error": str(e)}
    
@app.post("/api/debug/test-with-verified-email")
async def test_with_verified_email():
    """Test con email FROM verificado en SendGrid"""
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        
        # Usar un dominio verificado en SendGrid
        message = Mail(
            from_email='vcapost23@gmail.com',  # Cambia esto
            to_emails='vcontrerasalcu@gmail.com',
            subject='TEST con Email Verificado',
            html_content='<h1>Test con email verificado</h1>'
        )
        
        sg = SendGridAPIClient("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")
        response = sg.send(message)
        
        return {"status": response.status_code, "success": response.status_code == 202}
        
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)