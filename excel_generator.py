import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def generate_excel(processed_data_list):
    """
    Genera MÚLTIPLES archivos Excel agrupados por empresa
    Devuelve una lista de diccionarios con {empresa: nombre, archivo: excel_data, resumen: datos}
    """
    try:
        logger.info(f"📊 INICIANDO GENERACIÓN DE EXCEL POR EMPRESA")
        logger.info(f"📦 Total de elementos recibidos: {len(processed_data_list)}")
        
        if not processed_data_list:
            logger.error("❌ No hay datos para generar Excel")
            return []

        # 1. AGRUPAR POR EMPRESA con manejo de valores None
        empresas = {}
        elementos_sin_empresa = []
        
        for data in processed_data_list:
            # ✅ MANEJAR VALORES None EN VendorName
            empresa_nombre = data.get('VendorName')
            archivo_origen = data.get('archivo_origen', 'Desconocido')
            
            # ✅ SI NO HAY NOMBRE DE EMPRESA, USAR UN VALOR POR DEFECTO
            if not empresa_nombre:
                empresa_nombre = f"Empresa_No_Identificada_{hash(archivo_origen) % 10000:04d}"
                logger.warning(f"⚠️ Factura sin nombre de empresa: {archivo_origen}. Usando: {empresa_nombre}")
                elementos_sin_empresa.append(archivo_origen)
            
            if empresa_nombre not in empresas:
                empresas[empresa_nombre] = []
            
            # Agregar el archivo de origen a los datos para referencia
            data['archivo_origen'] = archivo_origen
            empresas[empresa_nombre].append(data)
        
        logger.info(f"🏢 Empresas detectadas: {len(empresas)}")
        if elementos_sin_empresa:
            logger.info(f"   ⚠️ Facturas sin empresa identificada: {len(elementos_sin_empresa)}")
        
        for empresa, datos in empresas.items():
            logger.info(f"   📋 {empresa}: {len(datos)} facturas")

        # 2. GENERAR UN EXCEL POR EMPRESA
        archivos_empresas = []
        
        for empresa_nombre, facturas_empresa in empresas.items():
            logger.info(f"📊 Generando Excel para: {empresa_nombre}")
            
            # ✅ VALIDAR QUE HAY DATOS VÁLIDOS
            if not facturas_empresa or len(facturas_empresa) == 0:
                logger.warning(f"⚠️ Saltando empresa {empresa_nombre}: sin facturas válidas")
                continue
            
            # Crear Excel para esta empresa
            excel_data = generar_excel_empresa(empresa_nombre, facturas_empresa)
            
            if excel_data:
                # Calcular resumen de IVA para esta empresa
                resumen_iva = calcular_resumen_iva_empresa(facturas_empresa)
                
                archivos_empresas.append({
                    'empresa': empresa_nombre,
                    'archivo': excel_data,
                    'cantidad_facturas': len(facturas_empresa),
                    'resumen_iva': resumen_iva
                })
                logger.info(f"✅ Excel generado para {empresa_nombre}")
            else:
                logger.error(f"❌ Error generando Excel para {empresa_nombre}")
        
        logger.info(f"✅ Generados {len(archivos_empresas)} archivos Excel")
        return archivos_empresas
        
    except Exception as e:
        logger.error(f"❌ Error generando Excel: {e}")
        
        # Crear un Excel de error como fallback
        try:
            error_workbook = Workbook()
            error_sheet = error_workbook.active
            error_sheet.title = "Error"
            error_sheet.append(['Error al generar el reporte'])
            error_sheet.append([f'Detalle: {str(e)}'])
            error_sheet.append([f'Datos recibidos: {len(processed_data_list)} elementos'])
            
            # ✅ AGREGAR INFORMACIÓN DE DEBUG
            if processed_data_list:
                error_sheet.append([''])
                error_sheet.append(['Datos recibidos:'])
                for i, data in enumerate(processed_data_list):
                    vendor_name = data.get('VendorName', 'No identificado')
                    archivo = data.get('archivo_origen', 'Desconocido')
                    error_sheet.append([f'Elemento {i+1}: {vendor_name} - {archivo}'])
            
            error_output = BytesIO()
            error_workbook.save(error_output)
            error_output.seek(0)
            
            return [{
                'empresa': 'Error_Procesamiento',
                'archivo': error_output.getvalue(),
                'cantidad_facturas': 0,
                'resumen_iva': {}
            }]
        except Exception as fallback_error:
            logger.error(f"❌ Error incluso en fallback: {fallback_error}")
            return []

def generar_excel_empresa(empresa_nombre, facturas_empresa):
    """
    Genera un archivo Excel completo para una empresa específica con manejo mejorado de datos limitados
    """
    try:
        workbook = Workbook()
        
        # Eliminar hoja por defecto
        if workbook.sheetnames:
            workbook.remove(workbook.active)
        
        # Estilos mejorados - ✅ IMPORTACIONES CORREGIDAS
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_font = Font(bold=True, size=14, color="2E74B5")
        warning_font = Font(color="FF0000", italic=True, bold=True)
        success_font = Font(color="00B050", italic=True)
        normal_font = Font(size=10)
        
        # Bordes - ✅ DEFINIDOS CORRECTAMENTE
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        money_format = '#,##0.00€'
        date_format = 'dd/mm/yyyy'
        
        # ✅ VALIDAR Y LIMPIAR NOMBRE DE EMPRESA
        if not empresa_nombre or empresa_nombre == 'None':
            empresa_nombre = "Empresa No Identificada"
            logger.warning(f"⚠️ Usando nombre por defecto para empresa: {empresa_nombre}")
        else:
            # Limpiar nombre para seguridad
            empresa_nombre = "".join(c for c in empresa_nombre if c.isalnum() or c in (' ', '-', '_', '.', '&'))
        
        # CONTADORES PARA ESTADÍSTICAS
        total_facturas = len(facturas_empresa)
        facturas_con_datos_completos = 0
        facturas_con_datos_limitiados = 0
        
        # Crear una hoja por cada factura de esta empresa
        for i, factura_data in enumerate(facturas_empresa):
            archivo_origen = factura_data.get('archivo_origen', f'Factura_{i+1}')
            
            # ✅ NOMBRE DE HOJA SEGURO (evitar caracteres inválidos)
            sheet_name = f"Factura_{i+1}"
            try:
                # Intentar usar nombre del archivo si es válido
                safe_name = "".join(c for c in archivo_origen if c.isalnum() or c in (' ', '-', '_'))
                if safe_name and len(safe_name) <= 28:  # Dejar margen para número
                    sheet_name = f"{safe_name}_{i+1}"
                else:
                    sheet_name = f"Factura_{i+1}"
            except:
                sheet_name = f"Factura_{i+1}"  # Usar nombre por defecto si hay error
            
            worksheet = workbook.create_sheet(title=sheet_name)
            current_row = 1
            
            # ✅ AGREGAR INDICADOR DE CALIDAD DE DATOS MEJORADO
            confidence_level = factura_data.get('confidence_level', 'unknown')
            procesamiento_tipo = factura_data.get('procesamiento', 'standard')
            es_datos_limitiados = confidence_level in ['low', 'basic'] or procesamiento_tipo == 'fallback_basico'
            
            if es_datos_limitiados:
                facturas_con_datos_limitiados += 1
                
                # BANNER DE ADVERTENCIA MEJORADO
                warning_cell = worksheet.cell(row=current_row, column=1, 
                                            value='⚠️ AVISO: DATOS LIMITADOS - VERIFICACIÓN MANUAL RECOMENDADA')
                worksheet.merge_cells(f'A{current_row}:D{current_row}')
                warning_cell.font = warning_font
                warning_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                warning_cell.alignment = Alignment(horizontal='center')
                current_row += 1
                
                worksheet.append(['Tipo de procesamiento:', 'Datos básicos (Azure no extrajo información completa)'])
                worksheet.append(['Archivo origen:', archivo_origen])
                
                if factura_data.get('error_original'):
                    worksheet.append(['Error en procesamiento:', factura_data.get('error_original')])
                    current_row += 1
                
                worksheet.append(['Recomendación:', 'Verificar manualmente los datos de la factura original'])
                current_row += 2  # Espacio adicional
                
            else:
                facturas_con_datos_completos += 1
                
                # INDICADOR DE ÉXITO PARA DATOS BUENOS
                success_cell = worksheet.cell(row=current_row, column=1, 
                                            value='✅ DATOS EXTRAÍDOS CORRECTAMENTE')
                worksheet.merge_cells(f'A{current_row}:D{current_row}')
                success_cell.font = success_font
                success_cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                success_cell.alignment = Alignment(horizontal='center')
                current_row += 1
                
                worksheet.append(['Tipo de procesamiento:', 'Extracción automática exitosa'])
                worksheet.append(['Nivel de confianza:', confidence_level.title() if confidence_level != 'unknown' else 'Alto'])
                current_row += 2

            # 1. INFORMACIÓN DE LA EMPRESA - SECCIÓN MEJORADA
            empresa_header = worksheet.cell(row=current_row, column=1, value='INFORMACIÓN DE LA EMPRESA')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            empresa_header.font = header_font
            empresa_header.fill = header_fill
            empresa_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # ✅ DATOS CON VALORES POR DEFECTO Y INDICADORES DE CALIDAD
            vendor_name = factura_data.get('VendorName', 'No identificado')
            vendor_tax_id = factura_data.get('VendorTaxId', 'No disponible')
            vendor_address = factura_data.get('VendorAddress', 'No disponible')
            
            # Marcar datos estimados con indicador visual
            if es_datos_limitiados and (vendor_name.startswith('Empresa_Desde_') or vendor_name == 'Empresa No Identificada'):
                vendor_name_display = f"{vendor_name} ⚠️"
            else:
                vendor_name_display = vendor_name
            
            worksheet.append(['Empresa:', vendor_name_display, '', ''])
            worksheet.append(['CIF/NIF:', vendor_tax_id, '', ''])
            
            # Manejar dirección multilínea
            if vendor_address and len(vendor_address) > 50:
                # Dividir dirección larga en múltiples filas
                address_parts = [vendor_address[i:i+50] for i in range(0, len(vendor_address), 50)]
                worksheet.append(['Dirección:', address_parts[0], '', ''])
                for part in address_parts[1:]:
                    worksheet.append(['', part, '', ''])
                    current_row += 1
            else:
                worksheet.append(['Dirección:', vendor_address, '', ''])
            
            current_row += 4  # Espacio después de información de empresa

            # 2. INFORMACIÓN ESPECÍFICA DE LA FACTURA
            factura_header = worksheet.cell(row=current_row, column=1, value='INFORMACIÓN DE LA FACTURA')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            factura_header.font = header_font
            factura_header.fill = header_fill
            factura_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            invoice_id = factura_data.get('InvoiceId', 'Sin número')
            invoice_date = factura_data.get('InvoiceDate', 'No especificada')
            invoice_total = factura_data.get('InvoiceTotal', 0)
            due_date = factura_data.get('DueDate', 'No especificada')
            
            worksheet.append(['Archivo origen:', archivo_origen, '', ''])
            worksheet.append(['Número Factura:', invoice_id, '', ''])
            
            # Formatear fecha correctamente
            try:
                if invoice_date and invoice_date != 'No especificada' and isinstance(invoice_date, str):
                    if 'T' in invoice_date:  # Formato ISO
                        invoice_date_obj = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
                        invoice_date = invoice_date_obj.strftime('%d/%m/%Y')
                    elif '-' in invoice_date:  # Formato YYYY-MM-DD
                        invoice_date_obj = datetime.strptime(invoice_date, '%Y-%m-%d')
                        invoice_date = invoice_date_obj.strftime('%d/%m/%Y')
            except (ValueError, AttributeError) as e:
                logger.warning(f"Error formateando fecha {invoice_date}: {e}")
                # Mantener fecha original si hay error
            
            worksheet.append(['Fecha Factura:', invoice_date, '', ''])
            
            if due_date and due_date != 'No especificada':
                try:
                    if 'T' in due_date:
                        due_date_obj = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
                        due_date = due_date_obj.strftime('%d/%m/%Y')
                    elif '-' in due_date:
                        due_date_obj = datetime.strptime(due_date, '%Y-%m-%d')
                        due_date = due_date_obj.strftime('%d/%m/%Y')
                except (ValueError, AttributeError):
                    pass
                worksheet.append(['Fecha Vencimiento:', due_date, '', ''])
            
            current_row += 4  # Espacio después de información de factura

            # 3. ARTÍCULOS DE LA FACTURA - TABLA MEJORADA
            items_header = worksheet.cell(row=current_row, column=1, value='ARTÍCULOS FACTURADOS')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            items_header.font = header_font
            items_header.fill = header_fill
            items_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Encabezados de tabla
            headers = ['Artículo', 'Unidades', 'Precio Unitario', 'Precio Total']
            worksheet.append(headers)
            
            # Aplicar estilo a encabezados
            for col in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            items = factura_data.get('Items', [])
            if items:
                subtotal_items = 0
                for item in items:
                    # ✅ VALORES POR DEFECTO PARA ITEMS
                    description = item.get('Description', 'Sin descripción')
                    quantity = item.get('Quantity', 0)
                    unit_price = item.get('UnitPrice', 0)
                    amount = item.get('Amount', 0)
                    
                    # Calcular amount si no está presente
                    if amount == 0 and quantity != 0 and unit_price != 0:
                        amount = quantity * unit_price
                    
                    subtotal_items += amount
                    
                    worksheet.append([
                        description,
                        quantity,
                        unit_price,
                        amount
                    ])
                    
                    # Aplicar bordes y formato a la fila
                    for col in range(1, 5):
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.font = normal_font
                        
                        # Formato numérico para columnas de precios
                        if col in [3, 4]:  # Precio Unitario y Precio Total
                            cell.number_format = money_format
                        elif col == 2:  # Unidades
                            cell.number_format = '0.##'
                    
                    current_row += 1
                
                # SUBTOTAL ARTÍCULOS
                worksheet.append(['SUBTOTAL ARTÍCULOS:', '', '', subtotal_items])
                subtotal_cell = worksheet.cell(row=current_row, column=4)
                subtotal_cell.font = Font(bold=True)
                subtotal_cell.number_format = money_format
                current_row += 1
                
            else:
                # NO HAY ARTÍCULOS
                worksheet.append(['No se encontraron artículos en esta factura', '', '', ''])
                for col in range(1, 5):
                    worksheet.cell(row=current_row, column=col).border = thin_border
                current_row += 1
            
            current_row += 1  # Espacio

            # 4. DETALLE DE IMPUESTOS - TABLA MEJORADA
            taxes_header = worksheet.cell(row=current_row, column=1, value='DETALLE DE IMPUESTOS')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            taxes_header.font = header_font
            taxes_header.fill = header_fill
            taxes_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            worksheet.append(['Tipo de IVA', 'Importe', '', ''])
            for col in range(1, 3):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            tax_details = factura_data.get('TaxDetails', [])
            total_impuestos = 0
            
            if tax_details:
                for tax in tax_details:
                    # ✅ VALORES POR DEFECTO PARA IMPUESTOS
                    rate = tax.get('Rate', '0%')
                    amount = tax.get('Amount', 0)
                    
                    total_impuestos += amount
                    
                    worksheet.append([
                        rate,
                        amount,
                        '', ''
                    ])
                    
                    # Aplicar bordes y formato
                    for col in range(1, 3):
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.font = normal_font
                        if col == 2:  # Columna de importe
                            cell.number_format = money_format
                    
                    current_row += 1
            else:
                worksheet.append(['No se encontraron impuestos', '', '', ''])
                for col in range(1, 3):
                    worksheet.cell(row=current_row, column=col).border = thin_border
                current_row += 1
            
            # TOTAL IMPUESTOS
            if tax_details:
                worksheet.append(['TOTAL IMPUESTOS:', total_impuestos, '', ''])
                tax_total_cell = worksheet.cell(row=current_row, column=2)
                tax_total_cell.font = Font(bold=True)
                tax_total_cell.number_format = money_format
                current_row += 1
            
            current_row += 1  # Espacio

            # 5. TOTAL FINAL DE LA FACTURA
            worksheet.append(['TOTAL FACTURA:', '', '', invoice_total])
            total_label_cell = worksheet.cell(row=current_row, column=1)
            total_value_cell = worksheet.cell(row=current_row, column=4)
            
            total_label_cell.font = total_font
            total_value_cell.font = total_font
            total_value_cell.number_format = money_format
            
            # Resaltar celda de total
            for col in range(1, 5):
                cell = worksheet.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = thin_border

            # 6. METADATOS ADICIONALES (solo si hay datos limitados)
            if es_datos_limitiados:
                current_row += 2
                worksheet.append(['INFORMACIÓN ADICIONAL DEL PROCESAMIENTO:', '', '', ''])
                worksheet.append(['Timestamp:', factura_data.get('timestamp_procesamiento', 'No disponible'), '', ''])
                worksheet.append(['Tipo archivo:', factura_data.get('tipo_archivo', 'No disponible'), '', ''])
                worksheet.append(['Índice procesamiento:', factura_data.get('indice_procesamiento', 'N/A'), '', ''])

            # AJUSTAR ANCHOS DE COLUMNAS MEJORADO
            column_widths = [45, 15, 18, 18]  # Ajustado para mejor visualización
            for col_idx, width in enumerate(column_widths, 1):
                column_letter = chr(64 + col_idx)  # A, B, C, D
                worksheet.column_dimensions[column_letter].width = width
            
            # CONGELAR PANELES (header visible al desplazar)
            worksheet.freeze_panes = 'A2'
            
            current_row += 2
        
        # 7. HOJA DE RESUMEN GENERAL MEJORADA
        resumen_sheet = workbook.create_sheet(title="📊 RESUMEN EMPRESA")
        current_row = 1
        
        # TÍTULO PRINCIPAL
        title_cell = resumen_sheet.cell(row=current_row, column=1, 
                                      value=f'RESUMEN GENERAL - {empresa_nombre.upper()}')
        resumen_sheet.merge_cells(f'A{current_row}:C{current_row}')
        title_cell.font = Font(bold=True, size=16, color="2E74B5")
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # ESTADÍSTICAS RÁPIDAS
        resumen_sheet.append(['ESTADÍSTICAS DE PROCESAMIENTO:', '', ''])
        resumen_sheet.merge_cells(f'A{current_row}:C{current_row}')
        resumen_sheet.cell(row=current_row, column=1).font = header_font
        resumen_sheet.cell(row=current_row, column=1).fill = header_fill
        current_row += 1
        
        resumen_sheet.append(['Total de facturas procesadas:', total_facturas, ''])
        resumen_sheet.append(['Facturas con datos completos:', facturas_con_datos_completos, '✅'])
        resumen_sheet.append(['Facturas con datos limitados:', facturas_con_datos_limitiados, '⚠️'])
        
        # Calcular porcentaje de éxito
        if total_facturas > 0:
            porcentaje_exito = (facturas_con_datos_completos / total_facturas) * 100
            resumen_sheet.append(['Tasa de éxito:', f'{porcentaje_exito:.1f}%', ''])
        
        current_row += 2
        
        # LISTA DETALLADA DE FACTURAS
        resumen_sheet.append(['LISTA DETALLADA DE FACTURAS:', '', ''])
        resumen_sheet.merge_cells(f'A{current_row}:C{current_row}')
        resumen_sheet.cell(row=current_row, column=1).font = header_font
        resumen_sheet.cell(row=current_row, column=1).fill = header_fill
        current_row += 1
        
        resumen_sheet.append(['Número Factura', 'Fecha', 'Total', 'Calidad'])
        for col in range(1, 5):
            resumen_sheet.cell(row=current_row, column=col).font = Font(bold=True)
            resumen_sheet.cell(row=current_row, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        current_row += 1
        
        total_general = 0
        for factura in facturas_empresa:
            invoice_id = factura.get('InvoiceId', 'Sin número')
            invoice_date = factura.get('InvoiceDate', 'No especificada')
            invoice_total = factura.get('InvoiceTotal', 0)
            confidence = factura.get('confidence_level', 'unknown')
            
            # Formatear fecha para resumen
            try:
                if invoice_date and invoice_date != 'No especificada' and isinstance(invoice_date, str):
                    if 'T' in invoice_date:
                        invoice_date_obj = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
                        invoice_date = invoice_date_obj.strftime('%d/%m/%Y')
            except:
                pass
            
            # Indicador de calidad
            if confidence in ['enhanced', 'high']:
                calidad = '✅ Alta'
            elif confidence in ['basic', 'low']:
                calidad = '⚠️ Básica'
            else:
                calidad = '🔍 Standard'
            
            resumen_sheet.append([invoice_id, invoice_date, invoice_total, calidad])
            total_general += invoice_total if invoice_total else 0
            current_row += 1
        
        current_row += 1
        
        # DETALLE DE IVA POR TIPO (AGREGADO)
        resumen_iva = calcular_resumen_iva_empresa(facturas_empresa)
        if resumen_iva:
            resumen_sheet.append(['DETALLE DE IVA POR TIPO:', '', ''])
            resumen_sheet.merge_cells(f'A{current_row}:C{current_row}')
            resumen_sheet.cell(row=current_row, column=1).font = header_font
            resumen_sheet.cell(row=current_row, column=1).fill = header_fill
            current_row += 1
            
            resumen_sheet.append(['Tipo de IVA', 'Total Importe', ''])
            for col in range(1, 3):
                resumen_sheet.cell(row=current_row, column=col).font = Font(bold=True)
                resumen_sheet.cell(row=current_row, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            current_row += 1
            
            for tipo_iva, importe in resumen_iva.items():
                resumen_sheet.append([tipo_iva, importe, ''])
                resumen_sheet.cell(row=current_row, column=2).number_format = money_format
                current_row += 1
        
        # TOTAL GENERAL
        current_row += 1
        resumen_sheet.append(['TOTAL GENERAL EMPRESA:', total_general, ''])
        resumen_sheet.cell(row=current_row, column=1).font = total_font
        resumen_sheet.cell(row=current_row, column=2).font = total_font
        resumen_sheet.cell(row=current_row, column=2).number_format = money_format
        
        # Ajustar anchos de columnas en resumen
        resumen_sheet.column_dimensions['A'].width = 35
        resumen_sheet.column_dimensions['B'].width = 20
        resumen_sheet.column_dimensions['C'].width = 15
        resumen_sheet.column_dimensions['D'].width = 12
        
        # CONGELAR PANELES EN RESUMEN
        resumen_sheet.freeze_panes = 'A3'
        
        # Guardar en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"✅ Excel generado para {empresa_nombre}: {total_facturas} facturas "
                   f"({facturas_con_datos_completos} completas, {facturas_con_datos_limitiados} limitadas)")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"❌ Error generando Excel para {empresa_nombre}: {e}")
        
        # ✅ FALLBACK MEJORADO: Crear Excel simple con información básica
        try:
            fallback_workbook = Workbook()
            fallback_sheet = fallback_workbook.active
            fallback_sheet.title = "Resumen Básico"
            
            fallback_sheet.append(['RESUMEN DE FACTURAS - ' + empresa_nombre])
            fallback_sheet.append([f'Error durante generación: {str(e)}'])
            fallback_sheet.append([''])
            fallback_sheet.append(['Facturas procesadas:', len(facturas_empresa)])
            fallback_sheet.append([''])
            
            # Información básica de cada factura
            fallback_sheet.append(['Detalle de facturas procesadas:'])
            fallback_sheet.append(['Archivo', 'Empresa', 'Número', 'Total'])
            
            for i, factura in enumerate(facturas_empresa):
                fallback_sheet.append([
                    factura.get('archivo_origen', f'Factura_{i+1}'),
                    factura.get('VendorName', 'No identificado'),
                    factura.get('InvoiceId', 'Sin número'),
                    factura.get('InvoiceTotal', 0)
                ])
            
            fallback_output = BytesIO()
            fallback_workbook.save(fallback_output)
            fallback_output.seek(0)
            
            logger.info(f"✅ Fallback Excel generado para {empresa_nombre}")
            return fallback_output.getvalue()
            
        except Exception as fallback_error:
            logger.error(f"❌ Error incluso en fallback: {fallback_error}")
            return None

def calcular_resumen_iva_empresa(facturas_empresa):
    """
    Calcula el total de IVA por tipo para todas las facturas de una empresa
    """
    resumen_iva = {}
    
    for factura in facturas_empresa:
        tax_details = factura.get('TaxDetails', [])
        for tax in tax_details:
            tipo_iva = tax.get('Rate', '0%')
            importe = tax.get('Amount', 0)
            
            if tipo_iva not in resumen_iva:
                resumen_iva[tipo_iva] = 0
            resumen_iva[tipo_iva] += importe
    
    logger.info(f"📊 Resumen IVA para empresa: {resumen_iva}")
    return resumen_iva

# Función de compatibilidad (por si otros partes del código esperan la función antigua)
def generate_single_excel(processed_data_list):
    """
    Función de compatibilidad - genera un solo Excel como antes
    """
    logger.warning("⚠️ Usando función de compatibilidad - genera un solo Excel")
    
    # Usar la lógica original simplificada
    try:
        workbook = Workbook()
        if workbook.sheetnames:
            workbook.remove(workbook.active)
        
        # Crear una hoja simple
        worksheet = workbook.create_sheet(title="Facturas Consolidadas")
        worksheet.append(['ADVERTENCIA: Este es un Excel de compatibilidad'])
        worksheet.append(['Se recomienda usar la nueva función generate_excel()'])
        worksheet.append(['Número de facturas procesadas:', len(processed_data_list)])
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
    except Exception as e:
        logger.error(f"❌ Error en función de compatibilidad: {e}")
        return None