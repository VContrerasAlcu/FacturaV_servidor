# excel_generator_simple.py
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

def generate_simplified_excel(processed_data_list):
    """
    Genera Excel simplificado SOLO con campos esenciales
    """
    try:
        logger.info(f"Generando Excel simplificado para {len(processed_data_list)} elementos")
        
        if not processed_data_list:
            logger.error("No hay datos para generar Excel")
            return []

        # Agrupar por empresa
        empresas = {}
        
        for data in processed_data_list:
            empresa_nombre = data.get('VendorName', 'Empresa No Identificada')
            
            if not empresa_nombre or empresa_nombre == 'None':
                empresa_nombre = "Empresa No Identificada"
            
            if empresa_nombre not in empresas:
                empresas[empresa_nombre] = []
            
            empresas[empresa_nombre].append(data)
        
        logger.info(f"Empresas detectadas: {len(empresas)}")
        
        # Generar Excel por empresa
        archivos_empresas = []
        
        for empresa_nombre, facturas_empresa in empresas.items():
            logger.info(f"Generando Excel para: {empresa_nombre} ({len(facturas_empresa)} facturas)")
            
            excel_data = generar_excel_empresa_simplificado(empresa_nombre, facturas_empresa)
            
            if excel_data:
                # Calcular resumen
                total_facturas = len(facturas_empresa)
                total_importe = sum(factura.get('InvoiceTotal', 0) for factura in facturas_empresa)
                resumen_iva = calcular_resumen_iva_completo(facturas_empresa)
                
                archivos_empresas.append({
                    'empresa': empresa_nombre,
                    'archivo': excel_data,
                    'cantidad_facturas': total_facturas,
                    'total_importe': total_importe,
                    'resumen_iva': resumen_iva
                })
        
        logger.info(f"Generados {len(archivos_empresas)} archivos Excel simplificados")
        return archivos_empresas
        
    except Exception as e:
        logger.error(f"Error generando Excel simplificado: {e}")
        return []

def generar_excel_empresa_simplificado(empresa_nombre, facturas_empresa):
    """
    Genera Excel con UNA HOJA POR FACTURA + HOJA RESUMEN
    """
    try:
        workbook = Workbook()
        
        # Eliminar hoja por defecto
        if workbook.sheetnames:
            workbook.remove(workbook.active)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=11, color="2E74B5")
        total_font = Font(bold=True, size=12, color="2E74B5")
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # CREAR UNA HOJA POR CADA FACTURA
        for factura_idx, factura in enumerate(facturas_empresa):
            sheet_name = f"Factura_{factura_idx + 1}"
            if len(sheet_name) > 31:  # Limite Excel
                sheet_name = f"F_{factura_idx + 1}"
            
            worksheet = workbook.create_sheet(title=sheet_name)
            current_row = 1
            
            # TITULO DE LA FACTURA
            title_cell = worksheet.cell(row=current_row, column=1, 
                                      value=f'FACTURA {factura_idx + 1} - {empresa_nombre}')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            title_cell.font = Font(bold=True, size=14, color="2E74B5")
            title_cell.alignment = Alignment(horizontal='center')
            current_row += 2
            
            # INFORMACION DEL VENDEDOR
            vendor_header = worksheet.cell(row=current_row, column=1, value='INFORMACION DEL VENDEDOR')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            vendor_header.font = header_font
            vendor_header.fill = header_fill
            vendor_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            worksheet.append(['Empresa:', factura.get('VendorName', 'No identificado'), '', ''])
            worksheet.append(['CIF/NIF:', factura.get('VendorTaxId', 'No disponible'), '', ''])
            worksheet.append(['Direccion:', factura.get('VendorAddress', 'No disponible'), '', ''])
            current_row += 4
            
            # INFORMACION DE LA FACTURA
            invoice_header = worksheet.cell(row=current_row, column=1, value='INFORMACION DE LA FACTURA')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            invoice_header.font = header_font
            invoice_header.fill = header_fill
            invoice_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            worksheet.append(['Numero Factura:', factura.get('InvoiceId', 'Sin numero'), '', ''])
            worksheet.append(['Fecha Factura:', formatear_fecha(factura.get('InvoiceDate')), '', ''])
            worksheet.append(['Archivo Origen:', factura.get('archivo_origen', 'Desconocido'), '', ''])
            current_row += 3
            
            # DETALLE DE IMPUESTOS
            taxes_header = worksheet.cell(row=current_row, column=1, value='DETALLE DE IMPUESTOS')
            worksheet.merge_cells(f'A{current_row}:D{current_row}')
            taxes_header.font = header_font
            taxes_header.fill = header_fill
            taxes_header.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Encabezados tabla impuestos
            worksheet.append(['Tipo de IVA', 'Tasa', 'Importe', ''])
            for col in range(1, 4):
                cell = worksheet.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = thin_border
            current_row += 1
            
            tax_details = factura.get('TaxDetails', [])
            total_impuestos = 0
            
            if tax_details:
                for tax in tax_details:
                    rate = tax.get('Rate', '0%')
                    amount = tax.get('Amount', 0)
                    total_impuestos += amount
                    
                    worksheet.append([
                        'IVA',
                        rate,
                        amount,
                        ''
                    ])
                    
                    # Formato de la fila
                    for col in range(1, 4):
                        cell = worksheet.cell(row=current_row, column=col)
                        cell.border = thin_border
                        if col == 3:  # Columna importe
                            cell.number_format = '#,##0.00€'
                    
                    current_row += 1
            else:
                worksheet.append(['No se detectaron impuestos', '', '', ''])
                for col in range(1, 4):
                    worksheet.cell(row=current_row, column=col).border = thin_border
                current_row += 1
            
            # TOTALES
            current_row += 1
            invoice_total = factura.get('InvoiceTotal', 0)
            
            worksheet.append(['SUBTOTAL (sin impuestos):', '', invoice_total - total_impuestos, ''])
            worksheet.append(['TOTAL IMPUESTOS:', '', total_impuestos, ''])
            worksheet.append(['TOTAL FACTURA:', '', invoice_total, ''])
            
            # Formato totales
            for row_offset in range(3):
                for col in range(1, 4):
                    cell = worksheet.cell(row=current_row + row_offset, column=col)
                    cell.border = thin_border
                    if col == 3:  # Columna importe
                        cell.number_format = '#,##0.00€'
                        if row_offset == 2:  # Fila de total
                            cell.font = total_font
            
            # Ajustar anchos de columnas
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 10
            
            # Congelar paneles
            worksheet.freeze_panes = 'A2'
        
        # HOJA DE RESUMEN GENERAL
        resumen_sheet = workbook.create_sheet(title="RESUMEN GENERAL")
        current_row = 1
        
        # TITULO
        title_cell = resumen_sheet.cell(row=current_row, column=1, 
                                      value=f'RESUMEN GENERAL - {empresa_nombre.upper()}')
        resumen_sheet.merge_cells(f'A{current_row}:H{current_row}')
        title_cell.font = Font(bold=True, size=16, color="2E74B5")
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # ESTADISTICAS RAPIDAS
        total_facturas = len(facturas_empresa)
        total_importe = sum(f.get('InvoiceTotal', 0) for f in facturas_empresa)
        total_impuestos = sum(sum(tax.get('Amount', 0) for tax in f.get('TaxDetails', [])) for f in facturas_empresa)
        
        resumen_sheet.append(['ESTADISTICAS GENERALES:', '', '', '', '', '', '', ''])
        resumen_sheet.merge_cells(f'A{current_row}:H{current_row}')
        resumen_sheet.cell(row=current_row, column=1).font = header_font
        resumen_sheet.cell(row=current_row, column=1).fill = header_fill
        current_row += 1
        
        resumen_sheet.append([
            'Total Facturas:', total_facturas,
            'Total Importe:', f'€{total_importe:,.2f}',
            'Total Impuestos:', f'€{total_impuestos:,.2f}',
            'Subtotal:', f'€{total_importe - total_impuestos:,.2f}'
        ])
        current_row += 2
        
        # TABLA RESUMEN DE FACTURAS
        resumen_sheet.append(['DETALLE POR FACTURA:', '', '', '', '', '', '', ''])
        resumen_sheet.merge_cells(f'A{current_row}:H{current_row}')
        resumen_sheet.cell(row=current_row, column=1).font = header_font
        resumen_sheet.cell(row=current_row, column=1).fill = header_fill
        current_row += 1
        
        headers = ['N Factura', 'Fecha', 'CIF/NIF', 'Subtotal', 'Total IVA', 'TOTAL', 'Tipos IVA', 'Archivo']
        resumen_sheet.append(headers)
        
        # Estilo encabezados
        for col in range(1, 9):
            cell = resumen_sheet.cell(row=current_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # DATOS DE FACTURAS
        for factura in facturas_empresa:
            invoice_id = factura.get('InvoiceId', 'Sin numero')
            invoice_date = formatear_fecha(factura.get('InvoiceDate'))
            tax_id = factura.get('VendorTaxId', 'No disponible')
            invoice_total = factura.get('InvoiceTotal', 0)
            
            # Calcular subtotal e IVA
            tax_details = factura.get('TaxDetails', [])
            total_iva = sum(tax.get('Amount', 0) for tax in tax_details)
            subtotal = invoice_total - total_iva
            
            # Tipos de IVA utilizados
            tipos_iva = ", ".join([tax.get('Rate', 'N/A') for tax in tax_details]) if tax_details else "No IVA"
            
            fila = [
                invoice_id,
                invoice_date,
                tax_id,
                subtotal,
                total_iva,
                invoice_total,
                tipos_iva,
                factura.get('archivo_origen', 'Desconocido')
            ]
            
            resumen_sheet.append(fila)
            
            # Formato de la fila
            for col in range(1, 9):
                cell = resumen_sheet.cell(row=current_row, column=col)
                cell.border = thin_border
                
                # Formato numerico para columnas de dinero
                if col in [4, 5, 6]:
                    cell.number_format = '#,##0.00€'
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # TOTALES FINALES
        current_row += 1
        start_data_row = current_row - len(facturas_empresa) - 1
        
        resumen_sheet.append([
            'TOTALES:', '', '',
            f'=SUM(D{start_data_row}:D{current_row-1})',
            f'=SUM(E{start_data_row}:E{current_row-1})', 
            f'=SUM(F{start_data_row}:F{current_row-1})',
            '', ''
        ])
        
        # Formato totales
        for col in range(1, 9):
            cell = resumen_sheet.cell(row=current_row, column=col)
            cell.font = total_font
            cell.border = thin_border
            if col in [4, 5, 6]:
                cell.number_format = '#,##0.00€'
        
        # RESUMEN DE IVA POR TIPO
        current_row += 2
        resumen_iva = calcular_resumen_iva_completo(facturas_empresa)
        
        if resumen_iva:
            resumen_sheet.append(['RESUMEN DE IVA POR TIPO:', '', '', '', '', '', '', ''])
            resumen_sheet.merge_cells(f'A{current_row}:H{current_row}')
            resumen_sheet.cell(row=current_row, column=1).font = header_font
            resumen_sheet.cell(row=current_row, column=1).fill = header_fill
            current_row += 1
            
            resumen_sheet.append(['Tipo de IVA', 'Total Importe', 'N Facturas', '', '', '', '', ''])
            for col in range(1, 4):
                cell = resumen_sheet.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = thin_border
            current_row += 1
            
            for tipo_iva, importe in resumen_iva.items():
                # Contar facturas con este tipo de IVA
                num_facturas = sum(1 for f in facturas_empresa 
                                 if any(tax.get('Rate') == tipo_iva for tax in f.get('TaxDetails', [])))
                
                resumen_sheet.append([tipo_iva, importe, num_facturas, '', '', '', '', ''])
                resumen_sheet.cell(row=current_row, column=2).number_format = '#,##0.00€'
                current_row += 1
        
        # Ajustar anchos de columnas en resumen
        column_widths = [20, 12, 15, 12, 12, 15, 20, 25]
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = chr(64 + col_idx)
            resumen_sheet.column_dimensions[col_letter].width = width
        
        # Congelar paneles
        resumen_sheet.freeze_panes = 'A3'
        
        # Guardar en memoria
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"Excel simplificado generado para {empresa_nombre}")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generando Excel para {empresa_nombre}: {e}")
        return None

def formatear_fecha(fecha):
    """Formatea fecha para Excel"""
    if not fecha:
        return 'No especificada'
    
    try:
        if isinstance(fecha, str):
            if 'T' in fecha:
                fecha_obj = datetime.fromisoformat(fecha.replace('Z', '+00:00'))
                return fecha_obj.strftime('%d/%m/%Y')
            elif '-' in fecha:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
                return fecha_obj.strftime('%d/%m/%Y')
        return str(fecha)
    except:
        return str(fecha)

def calcular_resumen_iva_completo(facturas_empresa):
    """Calcula resumen completo de IVA por tipo"""
    resumen_iva = {}
    
    for factura in facturas_empresa:
        tax_details = factura.get('TaxDetails', [])
        for tax in tax_details:
            tipo_iva = tax.get('Rate', '0%')
            importe = tax.get('Amount', 0)
            
            if tipo_iva not in resumen_iva:
                resumen_iva[tipo_iva] = 0
            resumen_iva[tipo_iva] += importe
    
    return resumen_iva